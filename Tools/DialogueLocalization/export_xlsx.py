"""Create a translator workbook from C++-joined manifest/PO rows."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from common import ENTRY_COLUMNS, as_text, export_fingerprint, new_workbook_id, utc_now


SPEAKER_COLUMNS = (
    "SpeakerId", "DisplayName", "SourceText", "SourceHash", "TextTableId", "TextTableKey",
    "LocNamespace", "LocKey", "MsgCtxt", "MsgId", "Translation", "TranslatorNote",
)


def write_table(sheet, headers, rows):
    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3F5F7F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 2, 12), 28)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--culture", required=True)
    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text(encoding="utf-8"))
    rows = data.get("rows", [])
    target = data.get("localizationTarget", "LostRunic")
    schema_version = int(data.get("schemaVersion", 1))
    workbook_id = new_workbook_id()
    speaker_rows = data.get("speakers", [])
    fingerprint = export_fingerprint(rows, schema_version, target, args.culture, speaker_rows)

    workbook = Workbook()
    meta = workbook.active
    meta.title = "Meta"
    for row in (
        ("WorkbookId", workbook_id), ("SchemaVersion", schema_version), ("LocalizationTarget", target),
        ("TargetCulture", args.culture), ("ManifestHash", data.get("manifestHash", "")),
        ("PORevision", data.get("poRevision", "")), ("ExportFingerprint", fingerprint), ("ExportTime", utc_now()),
        ("TranslationPolicy", "Only Translation/TranslatorNote may be edited by translators."),
    ):
        meta.append(row)
    meta.column_dimensions["A"].width = 24
    meta.column_dimensions["B"].width = 80

    entries_sheet = workbook.create_sheet("Entries")
    key_map = {
        "ScriptId": "scriptId", "StringKey": "stringKey", "DisplayId": "displayId", "SourceText": "sourceText",
        "SourceHash": "sourceHash", "TextTableId": "textTableId", "TextTableKey": "textTableKey",
        "LocNamespace": "locNamespace", "LocKey": "locKey", "MsgCtxt": "msgCtxt", "MsgId": "msgId",
        "SpeakerId": "speakerId", "EntryType": "entryType", "SourceLine": "sourceLine", "Order": "order",
        "ChoicePath": "choicePath", "ConditionalPath": "conditionalPath", "FormatArgs": "formatArgs",
        "Translation": "translation", "TranslatorNote": "translatorNote", "TranslatorContext": "translatorContext",
    }
    entry_values = []
    for row in rows:
        values = []
        for column in ENTRY_COLUMNS:
            value = row.get(key_map[column], "")
            values.append(json.dumps(value, ensure_ascii=False) if column == "FormatArgs" else as_text(value))
        entry_values.append(tuple(values))
    write_table(entries_sheet, ENTRY_COLUMNS, entry_values)

    speakers = workbook.create_sheet("Speakers")
    write_table(speakers, SPEAKER_COLUMNS, [
        tuple(row.get(key, "") for key in (
            "speakerId", "displayName", "sourceText", "sourceHash", "textTableId", "textTableKey",
            "locNamespace", "locKey", "msgCtxt", "msgId", "translation", "translatorNote",
        ))
        for row in speaker_rows
    ])

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
