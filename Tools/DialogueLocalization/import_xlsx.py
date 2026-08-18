"""Convert a translator workbook into machine-only Import JSON."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from common import ENTRY_COLUMNS, export_fingerprint, parse_json_cell


def read_meta(sheet):
    return {str(row[0].value): row[1].value for row in sheet.iter_rows(min_row=1, max_col=2) if row[0].value}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--culture", required=True)
    args = parser.parse_args()

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    meta = read_meta(workbook["Meta"])
    if meta.get("TargetCulture") != args.culture:
        raise ValueError("Workbook target culture does not match the requested import culture")
    sheet = workbook["Entries"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if tuple(header) != ENTRY_COLUMNS:
        raise ValueError("Entries sheet columns do not match the locked schema")

    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(header, values))
        rows.append({
            "scriptId": row["ScriptId"], "stringKey": row["StringKey"], "sourceHash": row["SourceHash"],
            "textTableId": row["TextTableId"], "textTableKey": row["TextTableKey"],
            "locNamespace": row["LocNamespace"], "locKey": row["LocKey"], "msgCtxt": row["MsgCtxt"],
            "msgId": row["MsgId"], "sourceText": row["SourceText"],
            "formatArgs": parse_json_cell(row["FormatArgs"]), "translation": row["Translation"] or "",
            "translatorNote": row["TranslatorNote"] or "",
        })

    speakers_sheet = workbook["Speakers"]
    speaker_header = [cell.value for cell in next(speakers_sheet.iter_rows(min_row=1, max_row=1))]
    expected_speaker_columns = (
        "SpeakerId", "DisplayName", "SourceText", "SourceHash", "TextTableId", "TextTableKey",
        "LocNamespace", "LocKey", "MsgCtxt", "MsgId", "Translation", "TranslatorNote",
    )
    if tuple(speaker_header) != expected_speaker_columns:
        raise ValueError("Speakers sheet columns do not match the locked schema")
    speakers = []
    for values in speakers_sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(speaker_header, values))
        speakers.append({
            "speakerId": row["SpeakerId"], "displayName": row["DisplayName"],
            "sourceText": row["SourceText"], "sourceHash": row["SourceHash"],
            "textTableId": row["TextTableId"], "textTableKey": row["TextTableKey"],
            "locNamespace": row["LocNamespace"], "locKey": row["LocKey"],
            "msgCtxt": row["MsgCtxt"], "msgId": row["MsgId"],
            "translation": row["Translation"] or "", "translatorNote": row["TranslatorNote"] or "",
        })
    workbook.close()

    target = meta.get("LocalizationTarget", "LostRunic")
    schema_version = int(meta.get("SchemaVersion", 1))
    output = {
        "schemaVersion": schema_version, "workbookId": meta.get("WorkbookId", ""),
        "localizationTarget": target, "targetCulture": args.culture,
        "manifestHash": meta.get("ManifestHash", ""), "poRevision": meta.get("PORevision", ""),
        "workbookFingerprint": export_fingerprint(rows, schema_version, target, args.culture, speakers),
        "rows": rows, "speakers": speakers,
    }
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    Path(args.output).write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
